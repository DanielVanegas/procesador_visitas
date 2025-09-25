import os
import re
from docx import Document
from openpyxl import load_workbook
from datos import obtener_datos_visita, obtener_datos_espaciales
from config import PARAM_FILE_PATH
from utils import log
from PyPDF2 import PdfReader
from datetime import datetime

def generar_informe(consecutivo, carpeta_path):
    documentos_path = os.path.join(carpeta_path, 'Documentos')
    observaciones_path = os.path.join(documentos_path, "Observaciones.docx")

    if os.path.exists(observaciones_path):
        log(f"La carpeta {consecutivo} ya tiene informe generado. Se omite.", carpeta_path)
        return

    datos = obtener_datos_visita(consecutivo)
    if not datos:
        log(f"No hay datos para el consecutivo {consecutivo}.", carpeta_path)
        return

    lat = datos.get("latitud")
    lon = datos.get("longitud")

    if lat is None or lon is None:
        log(f"Coordenadas faltantes para {consecutivo}.", carpeta_path)
        return

    barrio, localidad, sitios, estrato = obtener_datos_espaciales(lat, lon)
    
    if barrio == "No encontrado" or localidad == "No encontrado":
        log(f"Datos espaciales no encontrados para {consecutivo}. Se genera Observaciones.docx. con datos parciales", carpeta_path)
    else:
        log(f"Datos espaciales completos para {consecutivo}.", carpeta_path)
        
    texto = f"""
Inmueble ubicado en la localidad de {localidad}, barrio {barrio}, al sur {datos['al_sur']}, norte {datos['al_norte']}, oriente {datos['al_oriente']}, occidente {datos['al_occidente']}. Tiene cercanía a {sitios}.

Sector {datos['sector']}

{datos['edificio']} ubicado en un lote {datos['lote']} de forma irregular, tipología {datos['tipologia']}, topografía {datos['topografia']}, construcción de {datos['construccion']} pisos, {datos['sotanos']} sótano(s), zonas comunes {datos['zonas_comunes']}.

Apartamento ubicado en {datos['piso']} piso, vista {datos['vista']}, diseño funcional, iluminación y ventilación natural, acabados normales.

""".strip()
    
    # --- Garajes dinámicos ---
    if datos.get("num_garajes"):
        unidades = "unidad" if datos["num_garajes"] == 1 else "unidades"

        # Generar la lista de "No. XX" según cantidad
        garajes_texto = ""
        if datos["num_garajes"] == 1:
            garajes_texto = "No. XX"
        else:
            numeros = [f"XX" for _ in range(datos["num_garajes"])]
            garajes_texto = "No. " + ", ".join(numeros[:-1]) + f" y {numeros[-1]}"

        texto += f"\n\nGaraje(s) {garajes_texto} ({datos['num_garajes']} {unidades})"

        if datos.get("garaje_tipo"):
            texto += f", {datos['garaje_tipo']}"
        if datos.get("garaje_cubierta"):
            texto += f", {datos['garaje_cubierta']}"
        texto += ". "

    # --- Depósitos dinámicos ---
    if datos.get("num_depositos"):
        unidades = "unidad" if datos["num_depositos"] == 1 else "unidades"

        # Generar la lista de "No. XX" según cantidad
        dp_texto = ""
        if datos["num_depositos"] == 1:
            dp_texto = "No. XX"
        else:
            numerosdp = [f"XX" for _ in range(datos["num_depositos"])]
            dp_texto = "No. " + ", ".join(numerosdp[:-1]) + f" y {numerosdp[-1]}"

        texto += f"Depósito(s) {dp_texto} ({datos['num_depositos']} {unidades})"
        texto += ".\n\n"

    # --- Condicion jurídica dinamica ---
    if datos.get("garaje_cond_juridica") == "privado" and datos.get("deposito_cond_juridica") == "privado":
        texto += "El(los) garaje(s) y el(los) depósito(s) es(son) de uso privado, según Escritura Pública suministrada."
    elif datos.get("garaje_cond_juridica") == "exclusivo" and datos.get("deposito_cond_juridica") == "privado":
        texto += "El(los) garaje(s) es(son) de uso exclusivo y el(los) depósito(s) de uso privado, según Escritura Pública suministrada."
    elif datos.get("garaje_cond_juridica") == "exclusivo" and datos.get("deposito_cond_juridica") == "exclusivo":
        texto += "El(los) garaje(s) y el(los) depósito(s) es(son) de uso exclusivo, según Escritura Pública suministrada."
    elif datos.get("garaje_cond_juridica") == "privado" and datos.get("deposito_cond_juridica") == "exclusivo":
        texto += "El(los) garaje(s) es(son) de uso privado y el(los) depósito(s) de uso exclusivo, según Escritura Pública suministrada."

    texto += "\n"

    # --- Áreas dinámicas ---
    if datos.get("area_construida") or datos.get("area_privada"):
        # Validar que area_construida exista y sea mayor que 0
        if datos.get("area_construida") is not None and float(datos["area_construida"]) > 0:
            texto += f"\n\nÁrea total construida de {datos['area_construida']} m²"

        # Validar que area_privada exista y sea mayor que 0
        if datos.get("area_privada") is not None and float(datos["area_privada"]) > 0:
            if "Área total construida" in texto:
                texto += f", área privada de {datos['area_privada']} m², tomadas de la documentación suministrada. Cuenta con el derecho de uso exclusivo sobre una zona común destinada a terraza con un área de 31.69 m², aproximadamente 28.61 m² medidos in situ."
            else:
                texto += f"\n\nÁrea privada de {datos['area_privada']} m², tomada de la documentación suministrada. Cuenta con el derecho de uso exclusivo sobre una zona común destinada a terraza con un área de 31.69 m², aproximadamente 28.61 m² medidos in situ."

        texto += "."

    # --- Fecha ---
    texto += f"\n\nFecha de la visita: {datos['fecha_formulario']}."

#Garaje(s) No. 39 y 40 (2 unidades), sencillo(s), cubierto(s), en servidumbre propia. Depósito(s) No. 206 (1 unidad). El(los) garaje(s) y el(los) depósito(s) son de uso privado, según Escritura Pública suministrada.

#Área total construida de 91.05 m², área privada de 85.82 m², cuenta con el derecho de uso exclusivo sobre una zona común destinada a terraza con un área de 31.69 m², aproximadamente 28.61 m² medidos in situ.

#Fecha de la visita: {datos['fecha_formulario']}.
#"""

    os.makedirs(documentos_path, exist_ok=True)
    doc = Document()
    doc.add_paragraph(texto.strip())
    doc.save(observaciones_path)
    log(f"Generado Observaciones.docx para {consecutivo}.", carpeta_path)

    param_dest_path = os.path.join(carpeta_path, os.path.basename(PARAM_FILE_PATH))
    if not os.path.exists(param_dest_path):
        log(f"No se encontró el archivo de parámetros en {consecutivo}.", carpeta_path)
        return

    wb = load_workbook(param_dest_path)
    ws = wb.active

    for archivo in os.listdir(documentos_path):
        if archivo.lower().startswith("formato") and archivo.lower().endswith(".pdf"):
            ruta_pdf = os.path.join(documentos_path, archivo)
            datos_pdf = extraer_datos_pdf(ruta_pdf)
            if datos_pdf:
                fecha_raw = datos_pdf.get("fecha_radicacion")
                if fecha_raw:
                    try:
                        ws["B2"] = datetime.strptime(fecha_raw, "%d/%m/%Y")
                        ws["B2"].number_format = "dd/mm/yyyy hh:mm"
                    except ValueError:
                        log(f"Formato de fecha inválido en PDF para {consecutivo}: {fecha_raw}", carpeta_path)
                        ws["B2"] = fecha_raw
                ws["C5"] = datos_pdf.get("nombre_solicitante", "")
                ws["D39"] = datos_pdf.get("valor_inmueble", "")
            break

    ws['C2'] = datos['fecha_formulario']
    ws['D2'] = datos['fecha_formulario']
    ws['C2'].number_format = "dd/mm/yyyy hh:mm"
    ws['D2'].number_format = "dd/mm/yyyy hh:mm"
    ws['H2'] = f"{consecutivo}-DVF"
    ws['C8'] = barrio.upper() if barrio != "No encontrado" else ""
    ws['E8'] = estrato if estrato != "No identificado" else ""
    ws['E9'] = localidad.upper() if localidad != "No encontrado" else ""
    ws['K105'] = f"{lat}, {lon}"
    ws['K103'] = f"https://www.google.com/maps?q={lat},{lon}"

    wb.save(param_dest_path)

    log(f"Generado informe y Excel para {consecutivo}.", carpeta_path)

def extraer_datos_pdf(ruta_pdf):
    if not os.path.exists(ruta_pdf):
        log(f"PDF no encontrado en {ruta_pdf}.", os.path.dirname(ruta_pdf))
        return None

    try:
        reader = PdfReader(ruta_pdf)
        texto = ""
        for page in reader.pages:
            texto += page.extract_text() + "\n"

        if not texto.strip():
            log(f"El PDF {ruta_pdf} no contiene texto extraíble. Posiblemente está escaneado.", os.path.dirname(ruta_pdf))
            return None
        
    except Exception as e:
        log(f"Error al leer el PDF {ruta_pdf}: {e}", os.path.dirname(ruta_pdf))
        return None

    datos = {}

    # Extraer Fecha de radicación
    match_fecha = re.search(r"Fecha de radicación\s+([0-9]{2}/[0-9]{2}/[0-9]{4})", texto)
    if match_fecha:
        datos["fecha_radicacion"] = match_fecha.group(1)

    # Extraer Nombre del solicitante
    match_nombre = re.search(r"Nombre y apellidos Solicitante\s+([A-Za-zÁÉÍÓÚÜÑáéíóúüñ'´`\- ]+)", texto)
    if match_nombre:
        datos["nombre_solicitante"] = match_nombre.group(1).title()

    # Extraer Valor del inmueble
    match_valor = re.search(r"Valor del Inmueble\s+(\d+)", texto)
    if match_valor:
        datos["valor_inmueble"] = int(match_valor.group(1))

    return datos